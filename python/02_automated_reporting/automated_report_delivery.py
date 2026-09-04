"""
============================================================
PROJECT: Automated Data Query & Report Delivery System
Author : Anil Kumar Nukala
Domain : Business Intelligence / Operations Reporting
Description:
    A configurable system that executes scheduled SQL
    queries against a database, formats the results into
    styled Excel reports, and delivers them via email or
    SFTP on a defined schedule.
    Supports multi-recipient routing, dynamic date params,
    and delivery confirmation logging.
============================================================
"""

import os
import io
import smtplib
import logging
import paramiko
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import pyodbc
import snowflake.connector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
)
log = logging.getLogger(__name__)


# ----------------------------------------------------------------
# Configuration dataclasses
# ----------------------------------------------------------------
@dataclass
class ReportConfig:
    report_name: str
    query: str
    schedule: str                    # cron expression for reference
    email_recipients: list[str]
    email_subject_template: str      # supports {date}, {report_name}
    email_body_template: str
    sftp_path: Optional[str] = None  # set to enable SFTP delivery
    sheet_name: str = "Data"
    freeze_panes: str = "A2"
    date_params: dict = field(default_factory=dict)  # injected at query time


# ----------------------------------------------------------------
# Report catalogue — add new reports here
# ----------------------------------------------------------------
REPORT_CATALOGUE: list[ReportConfig] = [
    ReportConfig(
        report_name="Daily Inventory Summary",
        query="""
            SELECT
                warehouse_name,
                category,
                COUNT(DISTINCT sku_id)          AS total_skus,
                SUM(quantity_on_hand)           AS total_qty_on_hand,
                SUM(quantity_on_hand * unit_cost) AS inventory_value_usd,
                SUM(CASE WHEN quantity_on_hand < reorder_point
                         THEN 1 ELSE 0 END)     AS skus_below_reorder
            FROM warehouse.inventory_snapshot
            WHERE snapshot_date = '{snapshot_date}'
            GROUP BY warehouse_name, category
            ORDER BY inventory_value_usd DESC
        """,
        schedule="0 7 * * 1-5",
        email_recipients=["procurement@company.com", "ops-manager@company.com"],
        email_subject_template="Daily Inventory Summary — {date}",
        email_body_template=(
            "Please find attached the Daily Inventory Summary for {date}.\n\n"
            "Highlights:\n"
            "• SKUs below reorder point are flagged in the report.\n"
            "• Data reflects end-of-day snapshot.\n\n"
            "Regards,\nData & Reporting Team"
        ),
        date_params={"snapshot_date": "today"},
    ),

    ReportConfig(
        report_name="Weekly Supplier Scorecard",
        query="""
            SELECT
                supplier_name,
                on_time_delivery_pct,
                fill_rate_pct,
                defect_rate_pct,
                composite_score,
                supplier_status
            FROM analytics.supplier_scorecard_weekly
            WHERE week_ending = '{week_ending}'
            ORDER BY composite_score DESC
        """,
        schedule="0 8 * * 1",
        email_recipients=["supply-chain@company.com"],
        email_subject_template="Weekly Supplier Scorecard — Week ending {date}",
        email_body_template=(
            "Attached is the supplier performance scorecard for the week ending {date}.\n\n"
            "Please review suppliers flagged as AT RISK.\n\n"
            "Regards,\nData & Reporting Team"
        ),
        date_params={"week_ending": "last_sunday"},
    ),
]


# ----------------------------------------------------------------
# Database query execution
# ----------------------------------------------------------------
def execute_query(
    query: str,
    date_params: dict,
    source: str = "snowflake",
) -> pd.DataFrame:
    """Execute a parameterized query and return a DataFrame."""
    today = date.today()

    # Resolve date parameter placeholders
    resolved_params = {}
    for key, val in date_params.items():
        if val == "today":
            resolved_params[key] = today.strftime("%Y-%m-%d")
        elif val == "last_sunday":
            days_back = (today.weekday() + 1) % 7
            resolved_params[key] = (today - __import__("datetime").timedelta(days=days_back)).strftime("%Y-%m-%d")
        else:
            resolved_params[key] = val

    rendered_query = query.format(**resolved_params)

    if source == "snowflake":
        conn = snowflake.connector.connect(
            account=os.getenv("SF_ACCOUNT"),
            user=os.getenv("SF_USER"),
            password=os.getenv("SF_PASSWORD"),
            warehouse=os.getenv("SF_WAREHOUSE"),
            database=os.getenv("SF_DATABASE"),
        )
    else:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={os.getenv('WMS_SERVER')};"
            f"DATABASE={os.getenv('WMS_DATABASE')};"
            f"UID={os.getenv('WMS_USER')};"
            f"PWD={os.getenv('WMS_PASSWORD')};"
        )
        conn = pyodbc.connect(conn_str)

    try:
        df = pd.read_sql(rendered_query, conn)
        log.info(f"Query returned {len(df):,} rows")
        return df
    finally:
        conn.close()


# ----------------------------------------------------------------
# Excel report builder with formatting
# ----------------------------------------------------------------
def build_styled_excel(df: pd.DataFrame, config: ReportConfig) -> bytes:
    """
    Write DataFrame to a formatted Excel workbook and
    return the raw bytes for attachment or SFTP upload.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=config.sheet_name, index=False)
        ws = writer.sheets[config.sheet_name]

        # ---- Header row styling ----
        header_fill  = PatternFill(fill_type="solid", fgColor="1F3864")   # dark navy
        header_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin = Side(style="thin", color="D0D0D0")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = cell_border

        # ---- Data row alternating fill ----
        light_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")
        data_font  = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = light_fill if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.font      = data_font
                cell.alignment = data_align
                cell.border    = cell_border

        # ---- Auto-fit column widths ----
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        # ---- Row height for header ----
        ws.row_dimensions[1].height = 30

        # ---- Freeze panes ----
        ws.freeze_panes = config.freeze_panes

        # ---- Auto filter ----
        ws.auto_filter.ref = ws.dimensions

    output.seek(0)
    return output.read()


# ----------------------------------------------------------------
# Email delivery
# ----------------------------------------------------------------
def send_email(
    recipients: list[str],
    subject: str,
    body: str,
    attachment_bytes: bytes,
    attachment_filename: str,
) -> None:
    """Send the report via SMTP with Excel attachment."""
    smtp_host = os.getenv("SMTP_HOST", "smtp.office365.com")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASSWORD")
    from_addr = os.getenv("SMTP_FROM", smtp_user)

    msg = MIMEMultipart()
    msg["From"]    = from_addr
    msg["To"]      = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{attachment_filename}"')
    msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(from_addr, recipients, msg.as_string())

    log.info(f"Email sent → {', '.join(recipients)}")


# ----------------------------------------------------------------
# SFTP delivery
# ----------------------------------------------------------------
def sftp_upload(file_bytes: bytes, remote_path: str) -> None:
    """Upload report bytes to a remote SFTP server."""
    sftp_host = os.getenv("SFTP_HOST")
    sftp_user = os.getenv("SFTP_USER")
    sftp_key  = os.getenv("SFTP_KEY_PATH")

    transport = paramiko.Transport((sftp_host, 22))
    transport.connect(username=sftp_user, pkey=paramiko.RSAKey.from_private_key_file(sftp_key))
    sftp = paramiko.SFTPClient.from_transport(transport)

    with sftp.file(remote_path, "wb") as f:
        f.write(file_bytes)

    sftp.close()
    transport.close()
    log.info(f"SFTP upload complete → {remote_path}")


# ----------------------------------------------------------------
# Delivery log
# ----------------------------------------------------------------
def log_delivery(
    report_name: str,
    status: str,
    rows: int,
    recipients: list[str],
    error: str = None,
) -> None:
    """Append a delivery record to the local audit log."""
    record = {
        "timestamp"  : datetime.utcnow().isoformat(),
        "report_name": report_name,
        "status"     : status,
        "rows"       : rows,
        "recipients" : ";".join(recipients),
        "error"      : error or "",
    }
    log_df = pd.DataFrame([record])
    log_path = "logs/delivery_log.csv"
    log_df.to_csv(log_path, mode="a", header=not os.path.exists(log_path), index=False)


# ----------------------------------------------------------------
# Main report runner
# ----------------------------------------------------------------
def run_report(config: ReportConfig) -> None:
    """Execute, format, and deliver a single report."""
    today_str = date.today().strftime("%Y-%m-%d")
    log.info(f"=== Starting report: {config.report_name} ===")

    try:
        # 1. Query
        df = execute_query(config.query, config.date_params)

        if df.empty:
            log.warning(f"No data returned for '{config.report_name}' — skipping delivery")
            log_delivery(config.report_name, "NO_DATA", 0, config.email_recipients)
            return

        # 2. Build Excel
        excel_bytes = build_styled_excel(df, config)
        filename = f"{config.report_name.replace(' ', '_')}_{today_str}.xlsx"

        # 3. Email
        subject = config.email_subject_template.format(
            date=today_str, report_name=config.report_name
        )
        body = config.email_body_template.format(
            date=today_str, report_name=config.report_name
        )
        send_email(config.email_recipients, subject, body, excel_bytes, filename)

        # 4. SFTP (optional)
        if config.sftp_path:
            remote_path = f"{config.sftp_path}/{filename}"
            sftp_upload(excel_bytes, remote_path)

        # 5. Log success
        log_delivery(config.report_name, "SUCCESS", len(df), config.email_recipients)
        log.info(f"Report '{config.report_name}' delivered — {len(df):,} rows")

    except Exception as exc:
        log.exception(f"Report '{config.report_name}' failed: {exc}")
        log_delivery(config.report_name, "FAILURE", 0, config.email_recipients, str(exc))
        raise


# ----------------------------------------------------------------
# Entry point — run all reports in catalogue
# ----------------------------------------------------------------
if __name__ == "__main__":
    os.makedirs("logs", exist_ok=True)

    for cfg in REPORT_CATALOGUE:
        try:
            run_report(cfg)
        except Exception:
            log.error(f"Skipping '{cfg.report_name}' after failure — continuing with next")
